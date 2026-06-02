#!/usr/bin/env python3
"""
merge_to_xlsx.py — Combine 3 CSV files per language folder into one XLSX with 3 tabs.
Adds accuracy, naturalness, notes columns at the end of each tab.
"""
import csv
from pathlib import Path
from openpyxl import Workbook

BASE = Path(__file__).parent.parent / "Test Case" / "Testing Output"

FILE_MAP = [
    {"glob": "*explain*.csv", "tab": "Explanations"},
    {"glob": "*sub-50*.csv", "tab": "Subtitles"},
    {"glob": "*ui*.csv", "tab": "UI"},
]

REVIEW_COLS = ["accuracy", "naturalness", "notes"]


def write_tab(ws, csv_path: Path):
    with open(csv_path, "r", encoding="utf-8-sig") as f:
        reader = csv.DictReader(f)
        fieldnames = list(reader.fieldnames)
        rows = list(reader)

    new_header = fieldnames + REVIEW_COLS

    for col_idx, header in enumerate(new_header, 1):
        ws.cell(row=1, column=col_idx, value=header)

    for row_idx, row in enumerate(rows, 2):
        for col_idx, col in enumerate(fieldnames, 1):
            ws.cell(row=row_idx, column=col_idx, value=row[col])

    for col_idx in range(1, len(new_header) + 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = 18


def select_directory() -> Path:
    """Interactively select a directory from BASE."""
    if not BASE.exists():
        print(f"Warning: {BASE} not found. Please enter a path manually.")
        path_str = input("Enter path to folder containing CSVs: ").strip()
        return Path(path_str).resolve()

    subdirs = [d for d in BASE.iterdir() if d.is_dir() and not d.name.startswith(".")]
    subdirs.sort()

    print("\nSelect a folder containing CSVs to merge:")
    for i, d in enumerate(subdirs):
        print(f"  {i+1}. {d.name}")
    print(f"  {len(subdirs)+1}. [Enter Custom Path]")

    choice = input(f"\nPick a number (1-{len(subdirs)+1}): ").strip()
    if not choice:
        return subdirs[0] if subdirs else BASE
    
    try:
        idx = int(choice) - 1
        if 0 <= idx < len(subdirs):
            return subdirs[idx]
        else:
            path_str = input("Enter custom path: ").strip()
            return Path(path_str).resolve()
    except ValueError:
        return Path(choice).resolve()


def merge_folder(folder: Path):
    if not folder.exists():
        print(f"Error: folder {folder} not found")
        return

    wb = Workbook()
    first = True
    found_any = False

    for file_spec in FILE_MAP:
        matches = list(folder.glob(file_spec["glob"]))
        if not matches:
            print(f"  Warning: no match for {file_spec['glob']} in {folder.name}")
            continue

        found_any = True
        csv_path = matches[0]
        if first:
            ws = wb.active
            ws.title = file_spec["tab"]
            first = False
        else:
            ws = wb.create_sheet(title=file_spec["tab"])

        print(f"  {folder.name}/{file_spec['tab']} <- {csv_path.name}")
        write_tab(ws, csv_path)

    if not found_any:
        print(f"No CSV matches found in {folder}. Skipping.")
        return

    output_name = f"{folder.name.lower()}_translations_review.xlsx"
    output_path = folder / output_name
    wb.save(output_path)
    print(f"Saved: {output_path}")


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Merge CSVs in a folder to a single XLSX review file.")
    parser.add_argument("--dir", help="Path to folder containing CSVs")
    parser.add_argument("--all", action="store_true", help="Process all subdirectories in BASE")
    args = parser.parse_args()

    if args.all:
        subdirs = [d for d in BASE.iterdir() if d.is_dir() and not d.name.startswith(".")]
        for d in subdirs:
            print(f"\nProcessing {d.name}...")
            merge_folder(d)
    elif args.dir:
        merge_folder(Path(args.dir).resolve())
    else:
        folder = select_directory()
        merge_folder(folder)


if __name__ == "__main__":
    main()
